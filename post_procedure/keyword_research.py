#!/usr/bin/env python3
"""
Keyword Research Tool: Yandex Wordstat + Google Trends
-------------------------------------------------------
Два источника данных:
  1. Yandex Wordstat (через cookies сессии — без браузера)
  2. Google Trends (pytrends — без авторизации)

Запуск:
  python3 keyword_research.py                        # только Google Trends
  python3 keyword_research.py --cookies SESSION_COOKIE  # + Wordstat

Как получить Wordstat cookie:
  1. Зайдите на wordstat.yandex.ru в Chrome/Firefox
  2. DevTools → Application → Cookies → wordstat.yandex.ru
  3. Скопируйте значение "Session_id"
  4. Запустите: python3 keyword_research.py --cookies "ЗНАЧЕНИЕ_SESSION_ID"
"""

import time
import json
import argparse
import sys
import re
from pathlib import Path
from datetime import datetime

import requests
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

# ── Keyword clusters ───────────────────────────────────────────────────────────

CLUSTERS = {
    "Post-Procedure: Лазер": [
        # Seed
        "что делать после лазерной шлифовки",
        "уход после лазера лица",
        "чем мазать после лазера",
        "крем после лазерной шлифовки",
        "реабилитация после лазера дома",
        "уход после лазерного пилинга",
        # Из suggest
        "что делать после лазерной шлифовки лица",
        "что нельзя делать после лазерной шлифовки лица",
        "что нельзя делать после лазерной шлифовки",
        "крем после лазерной шлифовки для лица",
        "крем мазь после лазерной шлифовки",
        "восстановление после лазерного пилинга",
        # Инсайт: аргосульфан = аптечный конкурент → SEO-возможность
        "аргосульфан после лазерной шлифовки чем заменить",
    ],
    "Post-Procedure: Биоревитализация": [
        # Seed
        "уход после биоревитализации",
        "что нельзя после биоревитализации",
        "крем после биоревитализации",
        "что можно наносить после биоревитализации",
        "уход за кожей после биоревитализации",
        # Из suggest
        "уход после биоревитализации лица",
        "уход после биоревитализации лица первый день",
        "уход после биоревитализации лица гиалуроновой кислотой",
        "что нельзя после биоревитализации лица делать",
        "крем после биоревитализации для лица заживляющий",
        "крем после биоревитализации для лица",
        "крем от синяков после биоревитализации",
        "уход за кожей лица после биоревитализации",
    ],
    "Post-Procedure: Пилинг и дермароллер": [
        # Seed
        "уход после химического пилинга",
        "что после пилинга нельзя",
        "чем увлажнять после пилинга",
        "уход после дермароллера",
        "что наносить после дермароллера",
        "можно ли крем после дермароллера",
        # Из suggest
        "после химического пилинга уход за кожей лица",
        "уход за кожей после химического пилинга",
        "что нельзя делать после пилинга лица у косметолога",
        "после пилинга лица что нельзя делать",
        "чем увлажнять кожу после пилинга",
        "чем увлажнять лицо после пилинга",
    ],
    "Post-Procedure: RF и аппаратная косметология": [
        # Seed
        "уход после RF лифтинга",
        "что после аппаратной косметологии",
        "уход после ультразвуковой чистки лица",
        "уход после микротоков",
        # Из suggest
        "уход после rf лифтинга игольчатого",
        "восстановление после rf лифтинга игольчатого",
        "уход после микроигольчатого рф лифтинга лица",
        # HIFU убран (suggest вернул нерелевантное — нет трафика)
        # Добавлены правильные альтернативы
        "уход после ультразвукового лифтинга лица",
        "что нельзя после SMAS лифтинга",
    ],
    "Мужской уход": [
        # Seed
        "уход за кожей лица мужчины",
        "крем для лица мужской профессиональный",
        "сыворотка для мужчин",
        "уход за кожей мужчины после 40",
        "мужская косметология дома",
        "профессиональная косметика для мужчин",
        "лучший крем для лица мужской",
        # Из suggest
        "уход за лицом мужчины после 40",
        "уход за лицом мужчины после 30",
        "сыворотка для мужчин для лица",
        "уход за лицом мужчины",
    ],
    "Scalp Care / Кожа головы": [
        "уход за кожей головы",
        "сыворотка для кожи головы",
        "пилинг для кожи головы",
        "выпадение волос лечение домашнее",
        "маска для кожи головы профессиональная",
        "как укрепить волосы дома профессиональными средствами",
        "уход за кожей головы профессиональные средства",
        "сыворотка от выпадения волос профессиональная",
    ],
    "Anti-age 60+": [
        "крем для лица после 60",
        "уход за кожей 60 лет",
        "антивозрастная косметика 60+",
        "крем для зрелой кожи профессиональный",
        "лучший крем для женщин после 60",
        "питательный крем для лица 60",
        "уход за лицом после 60 лет",
        "профессиональная косметика для зрелой кожи",
    ],
    "Подростковое акне": [
        "косметика от акне для подростков",
        "уход за подростковой кожей",
        "как убрать прыщи у подростка",
        "профессиональные средства от акне",
        "уход при подростковом акне",
        "крем от прыщей для подростков",
        "средства от акне для подростков",
        "уход за проблемной кожей подростка",
    ],
}

# ── Wordstat scraper ───────────────────────────────────────────────────────────

WORDSTAT_BASE = "https://wordstat.yandex.ru"
WORDSTAT_UA   = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)

# Кэш сессии — инициализируем один раз на все запросы
_ws_session: requests.Session | None = None
_ws_csrf:    str | None = None


def wordstat_init(session_id: str) -> bool:
    """
    Инициализация сессии Wordstat:
    1. GET главной страницы — получаем куки и CSRF-токен
    2. Возвращает True если авторизация успешна
    """
    global _ws_session, _ws_csrf
    _ws_session = requests.Session()
    _ws_session.headers.update({
        "User-Agent":      WORDSTAT_UA,
        "Accept-Language": "ru-RU,ru;q=0.9",
        "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    })
    _ws_session.cookies.set("Session_id", session_id, domain=".yandex.ru")

    try:
        r = _ws_session.get(f"{WORDSTAT_BASE}/", timeout=15)
        if r.status_code != 200:
            print(f"  ✗ Wordstat: страница вернула {r.status_code}")
            return False

        # Ищем CSRF-токен в нескольких местах
        import re
        csrf = None
        for pat in [
            r'"csrfToken"\s*:\s*"([^"]+)"',
            r'"csrf_token"\s*:\s*"([^"]+)"',
            r'data-csrf="([^"]+)"',
            r'name="_csrf"\s+content="([^"]+)"',
            r'"_csrf"\s*:\s*"([^"]+)"',
            r'csrf["\']?\s*:\s*["\']([a-zA-Z0-9_\-]+)["\']',
        ]:
            m = re.search(pat, r.text)
            if m:
                csrf = m.group(1)
                break

        # Также проверяем куки
        if not csrf:
            for ck in _ws_session.cookies:
                if "csrf" in ck.name.lower():
                    csrf = ck.value
                    break

        _ws_csrf = csrf

        # Проверяем авторизацию — ищем признаки залогиненного пользователя
        if "wordstat" in r.url and r.status_code == 200:
            print(f"  ✓ Wordstat сессия OK (CSRF: {'найден' if csrf else 'не найден'})")
            return True
        return False

    except Exception as e:
        print(f"  ✗ Wordstat init ошибка: {e}")
        return False


def wordstat_fetch(keyword: str, session_id: str) -> dict:
    """
    Получить показы из Yandex Wordstat.
    Пробуем несколько эндпоинтов — API менялся несколько раз.
    """
    global _ws_session, _ws_csrf

    if _ws_session is None:
        wordstat_init(session_id)

    sess = _ws_session

    # Общие заголовки для API-запросов
    api_headers = {
        "Content-Type":   "application/json;charset=UTF-8",
        "Accept":         "application/json, text/plain, */*",
        "Origin":         WORDSTAT_BASE,
        "Referer":        f"{WORDSTAT_BASE}/",
    }
    if _ws_csrf:
        api_headers["X-Csrf-Token"] = _ws_csrf

    payload = {
        "words":      keyword,
        "pageNumber": 1,
        "pageSize":   50,
        "regionIds":  [],
        "deviceTypeId": 0,
    }

    # Эндпоинты — перебираем по порядку
    endpoints = [
        ("POST", f"{WORDSTAT_BASE}/api/main/user/words/list"),
        ("POST", f"{WORDSTAT_BASE}/api/wordstat/stat"),
        ("GET",  f"{WORDSTAT_BASE}/wordstat/stat"),
    ]

    for method, url in endpoints:
        try:
            if method == "POST":
                r = sess.post(url, json=payload, headers=api_headers, timeout=15)
            else:
                r = sess.get(url, params={"words": keyword}, headers=api_headers, timeout=15)

            if r.status_code == 405:
                continue   # Метод не подходит — пробуем следующий
            if r.status_code == 401:
                print(f"  ✗ Wordstat: Session_id устарел или невалиден")
                return {"keyword": keyword, "shows": None, "source": "auth_error"}
            if r.status_code == 403:
                print(f"  ✗ Wordstat 403 — CSRF не принят, пробую переинициализацию")
                wordstat_init(session_id)
                continue
            if r.status_code != 200:
                continue

            # Пытаемся разобрать ответ
            try:
                data = r.json()
            except Exception:
                continue

            # Формат 1: {data: {words: [{keyword, shows}]}}
            items = (data.get("data") or {}).get("words") or []
            # Формат 2: {words: [...]}
            if not items:
                items = data.get("words") or []
            # Формат 3: плоский список
            if not items and isinstance(data, list):
                items = data

            for item in items:
                if isinstance(item, dict):
                    kw_field = item.get("keyword") or item.get("word") or ""
                    if kw_field.lower() == keyword.lower():
                        shows = item.get("shows") or item.get("count") or 0
                        return {"keyword": keyword, "shows": shows, "source": "wordstat"}

            if items:
                first = items[0] if isinstance(items[0], dict) else {}
                shows = first.get("shows") or first.get("count") or 0
                return {"keyword": keyword, "shows": shows, "source": "wordstat"}

        except Exception as e:
            pass   # Пробуем следующий эндпоинт

    # Ни один эндпоинт не сработал — пробуем HTML-парсинг
    return _wordstat_html_parse(keyword)


def _wordstat_html_parse(keyword: str) -> dict:
    """
    Запасной вариант: парсим HTML страницы результатов Wordstat.
    """
    if _ws_session is None:
        return {"keyword": keyword, "shows": None, "source": "wordstat_error"}
    try:
        r = _ws_session.get(
            f"{WORDSTAT_BASE}/",
            params={"words": keyword},
            headers={"Accept": "text/html"},
            timeout=15,
        )
        import re
        # Ищем число вида "1 234 показов" или "Показов в месяц: 1 234"
        patterns = [
            r'"shows"\s*:\s*(\d+)',
            r'(\d[\d\s]+)\s*показ',
            r'показов[^\d]*(\d[\d\s]+)',
        ]
        for pat in patterns:
            m = re.search(pat, r.text, re.IGNORECASE)
            if m:
                val = int(m.group(1).replace(" ", "").replace("\xa0", ""))
                return {"keyword": keyword, "shows": val, "source": "wordstat_html"}
    except Exception:
        pass
    return {"keyword": keyword, "shows": None, "source": "wordstat_error"}


# ── Yandex Suggest API ────────────────────────────────────────────────────────
# Публичное автодополнение Яндекса — без авторизации, без блокировок.
# Подсказки упорядочены по популярности → позиция = прокси частотности.

SUGGEST_URL = "https://suggest.yandex.ru/suggest-ya.cgi"
SUGGEST_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept-Language": "ru-RU,ru;q=0.9",
}

def suggest_fetch(keyword: str) -> list[str]:
    """Return Yandex autocomplete suggestions for keyword."""
    try:
        r = requests.get(
            SUGGEST_URL,
            params={"part": keyword, "uil": "ru", "v": "4", "lr": "213"},
            headers=SUGGEST_HEADERS,
            timeout=8,
        )
        if r.status_code == 200:
            data = r.json()
            # Format: [query, [suggestions], ...]
            if len(data) >= 2 and isinstance(data[1], list):
                return [s for s in data[1] if isinstance(s, str)]
    except Exception as e:
        print(f"  ✗ Suggest ошибка для '{keyword}': {e}")
    return []


def suggest_score(keyword: str, all_suggests: dict[str, list[str]]) -> int | None:
    """
    Оценка популярности 1–10 на основе позиции в подсказках.
    Логика: чем выше позиция подсказки — тем популярнее запрос.
    Возвращает 10 если запрос появляется на позиции 1, 1 если на позиции 10+.
    """
    kw_lower = keyword.lower()
    for seed, suggests in all_suggests.items():
        for i, s in enumerate(suggests):
            if kw_lower in s.lower() or s.lower() in kw_lower:
                score = max(1, 10 - i)
                return score
    return None


# ── Yandex SERP result count (конкуренция) ────────────────────────────────────

YANDEX_SEARCH = "https://yandex.ru/search/"
YANDEX_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Accept-Language": "ru-RU,ru;q=0.9",
    "Accept": "text/html,application/xhtml+xml",
}

def yandex_results_count(keyword: str) -> str:
    """
    Получить количество результатов из Яндекс-поиска.
    Возвращает строку вида '12 300' или '—'.
    """
    try:
        r = requests.get(
            YANDEX_SEARCH,
            params={"text": keyword, "lr": "213"},
            headers=YANDEX_HEADERS,
            timeout=10,
        )
        if r.status_code == 200:
            # Яндекс показывает "Нашлось N результатов" или "Результаты: N"
            import re
            patterns = [
                r'Нашлось\s+([\d\s]+)\s+(?:результат|ответ)',
                r'найдено\s+([\d\s]+)',
                r'"totalItems"\s*:\s*(\d+)',
                r'data-count="(\d+)"',
            ]
            for pat in patterns:
                m = re.search(pat, r.text, re.IGNORECASE)
                if m:
                    return m.group(1).replace("\xa0", " ").strip()
    except Exception:
        pass
    return "—"


def _suggest_relevant(seed: str, suggestion: str) -> bool:
    """
    Проверка релевантности подсказки: минимум 2 слова из seed должны
    присутствовать в suggestion (защита от мусора типа 'орхидея уход...')
    """
    seed_words = set(w for w in seed.lower().split() if len(w) > 3)
    sugg_words = suggestion.lower()
    matches = sum(1 for w in seed_words if w in sugg_words)
    return matches >= max(1, len(seed_words) // 2)


def fetch_all_suggests(keywords: list[str]) -> dict[str, list[str]]:
    """Fetch Yandex suggests for all keywords, filter irrelevant results."""
    result = {}
    for kw in keywords:
        raw      = suggest_fetch(kw)
        filtered = [s for s in raw if _suggest_relevant(kw, s)]
        result[kw] = filtered
        if raw and not filtered:
            # Все подсказки нерелевантны → запрос пустой в Яндексе
            result[kw] = []
        time.sleep(0.3)
    return result


# ── Scoring ────────────────────────────────────────────────────────────────────

def priority_score(wordstat_shows, suggest_score_val, competition) -> str:
    """
    Эвристика приоритета:
      HIGH   = есть спрос (wordstat или suggest) + низкая конкуренция
      MEDIUM = средний спрос или средняя конкуренция
      LOW    = проверить вручную
    """
    if wordstat_shows is not None:
        if wordstat_shows >= 5000 and competition != "Высокая":
            return "HIGH"
        elif wordstat_shows >= 1000:
            return "MEDIUM"
        else:
            return "LOW (проверить)"

    # Только suggest-скор
    score = suggest_score_val or 0
    if score >= 7 and competition == "Низкая":
        return "HIGH"
    elif score >= 5:
        return "MEDIUM"
    else:
        return "LOW (проверить)"


def competition_hint(keyword: str) -> str:
    """Грубая оценка конкуренции по признакам запроса."""
    kw = keyword.lower()
    # Транзакционные = высокая конкуренция
    if any(w in kw for w in ["купить", "цена", "магазин", "заказать"]):
        return "Высокая"
    # Информационные после-процедурные = низкая
    if any(w in kw for w in ["после", "как", "что", "чем", "можно", "нельзя"]):
        return "Низкая"
    return "Средняя"


def content_type(keyword: str) -> str:
    kw = keyword.lower()
    if any(w in kw for w in ["как", "что", "чем", "можно", "нельзя", "после"]):
        return "Статья (SEO)"
    if any(w in kw for w in ["купить", "цена", "заказать"]):
        return "Карточка товара"
    return "Статья / категория"


# ── Excel export ───────────────────────────────────────────────────────────────

BLUE_DARK  = "1A4A8A"
BLUE_MED   = "2E75B6"
BLUE_LIGHT = "BDD7EE"
ACCENT     = "1A6AB0"
GREEN_D    = "1E8B4C"
GREEN_L    = "C6EFCE"
RED_D      = "C0392B"
RED_L      = "FFC7CE"
YELLOW_L   = "FFEB9C"
GRAY_L     = "F2F2F2"
WHITE      = "FFFFFF"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=11, color="000000", name="Calibri") -> Font:
    return Font(bold=bold, size=size, color=color, name=name)


def _border() -> Border:
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def build_excel(rows: list[dict], out_path: str):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Сводная таблица ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Все кластеры"

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "Keyword Research — Профессиональная косметика B2C"
    title_cell.font = _font(bold=True, size=14, color=WHITE)
    title_cell.fill = _fill(BLUE_DARK)
    title_cell.alignment = _center()
    ws.row_dimensions[1].height = 30

    # Subtitle
    ws.merge_cells("A2:H2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Источник: Google Trends (RU) + Yandex Wordstat | {datetime.now().strftime('%d.%m.%Y')}"
    sub_cell.font = _font(size=10, color="595959")
    sub_cell.fill = _fill("E8F0FE")
    sub_cell.alignment = _left()
    ws.row_dimensions[2].height = 18

    # Header
    headers = [
        "Кластер", "Ключевой запрос",
        "Suggest score\n(1–10)", "Wordstat\n(показов/мес)",
        "Топ подсказка Яндекса", "Конкуренция",
        "Тип контента", "Приоритет", "Новые кандидаты из suggest",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = _font(bold=True, size=10, color=WHITE)
        cell.fill = _fill(BLUE_MED)
        cell.alignment = _center()
        cell.border = _border()
    ws.row_dimensions[3].height = 40

    # Data rows
    cluster_colors = {}
    palette = ["DCE6F1", "E2EFDA", "FFF2CC", "F4CCCC", "EAD1DC",
               "CFE2F3", "D9EAD3", "FCE5CD"]
    current_row = 4
    cluster_idx = 0
    prev_cluster = None

    for row in rows:
        cluster = row["cluster"]
        if cluster != prev_cluster:
            if cluster not in cluster_colors:
                cluster_colors[cluster] = palette[cluster_idx % len(palette)]
                cluster_idx += 1
            prev_cluster = cluster

        bg   = cluster_colors[cluster]
        prio = row["priority"]
        prio_color = GREEN_D if prio == "HIGH" else ("595959" if prio == "MEDIUM" else RED_D)
        prio_bg    = GREEN_L if prio == "HIGH" else (YELLOW_L if prio == "MEDIUM" else RED_L)
        comp       = row["competition"]
        comp_color = RED_D if comp == "Высокая" else (GREEN_D if comp == "Низкая" else "595959")

        sc  = row["suggest_score"] if row["suggest_score"] is not None else "—"
        ws_ = row["wordstat"]      if row["wordstat"]      is not None else "—"

        vals = [
            cluster,
            row["keyword"],
            sc,
            ws_,
            row["suggest_top"],
            comp,
            row["content_type"],
            prio,
            row["new_keywords"],
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.border = _border()
            cell.alignment = _left() if col in (1, 2, 5, 7, 9) else _center()
            cell.font = _font(size=10)
            if col == 1:
                cell.fill = _fill(bg)
                cell.font = _font(size=10, bold=True)
            elif col == 6:
                cell.font = _font(size=10, color=comp_color)
            elif col == 8:
                cell.fill = _fill(prio_bg)
                cell.font = _font(size=10, bold=True, color=prio_color)
            else:
                cell.fill = _fill(WHITE)

        ws.row_dimensions[current_row].height = 22
        current_row += 1

    # Column widths
    col_widths = [24, 46, 14, 16, 40, 14, 18, 14, 44]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:H{current_row - 1}"

    # ── Sheet 2: По кластерам ───────────────────────────────────────────────
    ws2 = wb.create_sheet("По кластерам")
    ws2.merge_cells("A1:F1")
    ws2["A1"].value = "Сводка по кластерам"
    ws2["A1"].font = _font(bold=True, size=13, color=WHITE)
    ws2["A1"].fill = _fill(BLUE_DARK)
    ws2["A1"].alignment = _center()
    ws2.row_dimensions[1].height = 28

    h2 = ["Кластер", "Кол-во запросов", "Avg Suggest score",
          "Макс Wordstat", "HIGH приоритет", "Рекомендация"]
    for col, h in enumerate(h2, 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = _font(bold=True, size=10, color=WHITE)
        cell.fill = _fill(BLUE_MED)
        cell.alignment = _center()
        cell.border = _border()
    ws2.row_dimensions[2].height = 30

    # Aggregate
    from collections import defaultdict
    by_cluster = defaultdict(list)
    for row in rows:
        by_cluster[row["cluster"]].append(row)

    r2 = 3
    for cluster, kws in by_cluster.items():
        sc_vals    = [k["suggest_score"] for k in kws if k["suggest_score"] is not None]
        ws_vals    = [k["wordstat"]      for k in kws if k["wordstat"]      is not None]
        high_count = sum(1 for k in kws if k["priority"] == "HIGH")
        avg_trends = round(sum(sc_vals) / len(sc_vals), 1) if sc_vals else None
        max_ws     = max(ws_vals) if ws_vals else None

        # Рекомендация
        if high_count >= 3:
            rec = "★ Приоритетный — начать немедленно"
        elif high_count >= 1:
            rec = "◎ Перспективный — план на мес 2"
        else:
            rec = "○ Проверить вручную в Wordstat"

        row_data = [
            cluster,
            len(kws),
            avg_trends if avg_trends is not None else "—",
            max_ws if max_ws is not None else "—",
            high_count,
            rec,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r2, column=col, value=val)
            cell.border = _border()
            cell.alignment = _left() if col in (1, 6) else _center()
            cell.font = _font(size=10, bold=(col == 1))
        r2 += 1

    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 18
    ws2.column_dimensions["E"].width = 18
    ws2.column_dimensions["F"].width = 38

    # ── Sheet 3: Инструкция ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Инструкция")
    ws3.column_dimensions["A"].width = 80
    instructions = [
        ("КАК ЧИТАТЬ ЭТОТ ФАЙЛ", True, 14, WHITE, BLUE_DARK),
        ("", False, 11, "000000", WHITE),
        ("Google Trends (0–100)", True, 12, WHITE, BLUE_MED),
        ("Относительный показатель интереса за 12 месяцев в России.", False, 11, "000000", GRAY_L),
        ("100 = пик популярности за период, 50 = вдвое меньше пика.", False, 11, "000000", WHITE),
        ("Значения > 30 — запрос стабильно ищут. < 10 — редкий хвост.", False, 11, "000000", GRAY_L),
        ("", False, 11, "000000", WHITE),
        ("Wordstat (показов/мес)", True, 12, WHITE, BLUE_MED),
        ("Реальные ежемесячные показы запроса по России.", False, 11, "000000", GRAY_L),
        ("< 500 — хвост (можно, но маленький объём).", False, 11, "000000", WHITE),
        ("500–5 000 — оптимально для SEO-статей без мощной конкуренции.", False, 11, "000000", GRAY_L),
        ("> 5 000 — высокий спрос, скорее всего высокая конкуренция.", False, 11, "000000", WHITE),
        ("", False, 11, "000000", WHITE),
        ("Приоритет", True, 12, WHITE, BLUE_MED),
        ("HIGH   — брать в работу первыми.", False, 11, "000000", GREEN_L),
        ("MEDIUM — следующий этап.", False, 11, "000000", YELLOW_L),
        ("LOW    — проверить вручную, возможно слишком узкий хвост.", False, 11, "000000", RED_L),
        ("", False, 11, "000000", WHITE),
        ("Как добавить Wordstat данные", True, 12, WHITE, BLUE_MED),
        ("1. Зайдите на wordstat.yandex.ru в браузере (залогиньтесь).", False, 11, "000000", GRAY_L),
        ("2. DevTools (F12) → Application → Cookies → wordstat.yandex.ru", False, 11, "000000", WHITE),
        ("3. Скопируйте значение 'Session_id'", False, 11, "000000", GRAY_L),
        ("4. Запустите скрипт:", False, 11, "000000", WHITE),
        ("   python3 keyword_research.py --cookies 'ВАШ_SESSION_ID'", False, 11, "595959", GRAY_L),
    ]
    for i, (text, bold, size, fcolor, bg) in enumerate(instructions, 1):
        cell = ws3.cell(row=i, column=1, value=text)
        cell.font = _font(bold=bold, size=size, color=fcolor)
        cell.fill = _fill(bg)
        cell.alignment = _left()
        ws3.row_dimensions[i].height = 22 if text else 8

    wb.save(out_path)
    print(f"\n✓ Excel сохранён: {out_path}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Keyword Research Tool")
    parser.add_argument(
        "--cookies", metavar="SESSION_ID",
        help="Yandex Wordstat Session_id cookie"
    )
    parser.add_argument(
        "--serp", action="store_true",
        help="Проверить конкуренцию через Yandex SERP (медленнее)"
    )
    parser.add_argument(
        "--debug-wordstat", action="store_true",
        help="Показать сырой ответ Wordstat для первого запроса"
    )
    parser.add_argument(
        "--out", default="keyword_report.xlsx",
        help="Путь к выходному Excel (по умолчанию: keyword_report.xlsx)"
    )
    args = parser.parse_args()

    out_path     = Path(args.out)
    use_wordstat = bool(args.cookies)

    all_keywords = [kw for kws in CLUSTERS.values() for kw in kws]

    print("=" * 60)
    print("  Keyword Research Tool — Профессиональная косметика B2C")
    print("=" * 60)
    print(f"  Кластеров:    {len(CLUSTERS)}")
    print(f"  Запросов:     {len(all_keywords)}")
    print(f"  Источник:     Yandex Suggest (автодополнение)")
    print(f"  Wordstat:     {'✓ реальные показы' if use_wordstat else '✗  добавьте --cookies'}")
    print(f"  SERP-конкур.: {'✓' if args.serp else '✗  добавьте --serp'}")
    print()

    # ── 1. Yandex Suggest ────────────────────────────────────────────────────
    print("▶ Получаю Yandex Suggest для всех запросов...")
    all_suggests = fetch_all_suggests(all_keywords)
    got = sum(1 for v in all_suggests.values() if v)
    print(f"  → Получены подсказки для {got}/{len(all_keywords)} запросов\n")

    # Expand: добавить подсказки как новые кандидаты (выводим, не добавляем в таблицу)
    new_kw_found = {}
    for kw, suggests in all_suggests.items():
        extras = [s for s in suggests if s.lower() != kw.lower()][:3]
        if extras:
            new_kw_found[kw] = extras

    # ── 2. Wordstat ──────────────────────────────────────────────────────────
    ws_data = {}
    if use_wordstat:
        print("▶ Инициализирую сессию Yandex Wordstat...")
        wordstat_init(args.cookies)

        if args.debug_wordstat:
            print("\n── DEBUG: пробую первый запрос напрямую ──")
            test_kw = all_keywords[0]
            for method in ["POST", "GET"]:
                for path in ["/api/main/user/words/list", "/api/wordstat/stat", "/wordstat/stat"]:
                    url = WORDSTAT_BASE + path
                    hdrs = {
                        "User-Agent": WORDSTAT_UA,
                        "Content-Type": "application/json",
                        "Accept": "application/json",
                        "Referer": WORDSTAT_BASE + "/",
                    }
                    if _ws_csrf:
                        hdrs["X-Csrf-Token"] = _ws_csrf
                    try:
                        if method == "POST":
                            resp = _ws_session.post(url, json={"words": test_kw, "pageNumber": 1, "pageSize": 10, "regionIds": [], "deviceTypeId": 0}, headers=hdrs, timeout=10)
                        else:
                            resp = _ws_session.get(url, params={"words": test_kw}, headers=hdrs, timeout=10)
                        print(f"  {method} {path} → {resp.status_code}")
                        if resp.status_code == 200:
                            print(f"  Ответ: {resp.text[:300]}")
                    except Exception as e:
                        print(f"  {method} {path} → ошибка: {e}")
            print("── END DEBUG ──\n")
        print("▶ Получаю показы...")
        for kw in all_keywords:
            result = wordstat_fetch(kw, args.cookies)
            ws_data[kw] = result.get("shows")
            val_str = f"{ws_data[kw]:,}".replace(",", " ") if ws_data[kw] else "—"
            sys.stdout.write(f"  → [{val_str:>10}] {kw}\n")
            time.sleep(1.5)
        got_ws = sum(1 for v in ws_data.values() if v is not None)
        print(f"  → Получено показов: {got_ws}/{len(all_keywords)}\n")
    else:
        for kw in all_keywords:
            ws_data[kw] = None

    # ── 3. SERP competition ──────────────────────────────────────────────────
    serp_data = {}
    if args.serp:
        print("▶ Проверяю конкуренцию через Yandex SERP...")
        for kw in all_keywords:
            count = yandex_results_count(kw)
            serp_data[kw] = count
            sys.stdout.write(f"  → [{count:>12}] {kw}\n")
            time.sleep(1.5)
        print()
    else:
        for kw in all_keywords:
            serp_data[kw] = "—"

    # ── 4. Assemble ──────────────────────────────────────────────────────────
    rows = []
    for cluster, keywords in CLUSTERS.items():
        for kw in keywords:
            suggests    = all_suggests.get(kw, [])
            sc          = suggest_score(kw, all_suggests)
            ws          = ws_data.get(kw)
            comp        = competition_hint(kw)
            prio        = priority_score(ws, sc, comp)
            rows.append({
                "cluster":      cluster,
                "keyword":      kw,
                "suggest_top":  suggests[0] if suggests else "—",
                "suggest_count":len(suggests),
                "suggest_score":sc,
                "wordstat":     ws,
                "serp_count":   serp_data.get(kw, "—"),
                "competition":  comp,
                "content_type": content_type(kw),
                "priority":     prio,
                "new_keywords": " | ".join(new_kw_found.get(kw, [])),
            })

    # ── 5. Console summary ───────────────────────────────────────────────────
    print("\n📊 Результаты:")
    print(f"  {'Запрос':<48} {'Suggest':>7}  {'WS':>8}  {'Приоритет'}")
    print("  " + "-" * 85)
    prev = None
    for row in rows:
        if row["cluster"] != prev:
            print(f"\n  ── {row['cluster']}")
            prev = row["cluster"]
        sc  = f"{row['suggest_score']:>2}" if row["suggest_score"] else " —"
        ws  = f"{row['wordstat']:>8}" if row["wordstat"] else "       —"
        print(f"  {row['keyword']:<48} {sc}  {ws}  {row['priority']}")

    # ── 6. Excel ─────────────────────────────────────────────────────────────
    build_excel(rows, str(out_path))

    # ── 7. Suggest expansion ─────────────────────────────────────────────────
    if new_kw_found:
        print("\n💡 Новые запросы из автодополнения (добавить в кластеры):")
        for seed, extras in list(new_kw_found.items())[:15]:
            print(f"  [{seed[:35]}] → {' | '.join(extras)}")

    high = [r for r in rows if r["priority"] == "HIGH"]
    med  = [r for r in rows if r["priority"] == "MEDIUM"]
    print(f"\n🔥 HIGH: {len(high)}  ◎ MEDIUM: {len(med)}  ○ LOW: {len(rows)-len(high)-len(med)}")
    for r in high:
        print(f"   ✓ [{r['cluster'][:28]}] {r['keyword']}")
    print()


if __name__ == "__main__":
    main()